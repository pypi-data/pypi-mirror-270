from typing import List

from openpyxl.worksheet.worksheet import Worksheet


class DictReader:

    def __init__(
        self,
        worksheet: Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        headers: List[str] = None,
        values_only=True,
        skip_blank_rows=False,
        strip_whitespace_in_headers=True,
    ):
        self.worksheet = worksheet
        self.min_row = min_row
        self.max_row = max_row
        self.min_col = min_col
        self.max_col = max_col
        self.values_only = values_only
        self.skip_blank_rows = skip_blank_rows
        self.strip_whitespace_in_headers = strip_whitespace_in_headers
        self.curr_row = 1

        if headers:
            self.headers = headers
            self.min_col = self.min_col or 1
            self.max_col = self.min_col + len(headers) - 1
        else:
            self.headers = self._read_headers()

    def _read_headers(self):
        headers = []
        row = self.min_row or 1
        min_col = self.min_col or 1
        max_col = self.max_col or self.worksheet.max_column
        for col in self.worksheet.iter_cols(
            min_row=row, max_row=row, min_col=min_col, max_col=max_col
        ):
            hv = col[0].value
            if not hv:
                break
            if self.strip_whitespace_in_headers:
                hv = hv.strip()
            if hv in headers:
                raise ValueError(f"Duplicate header found: {hv}: {headers}")
            headers.append(hv)
        if self.max_col is not None:
            if len(headers) != max_col + 1 - min_col:
                raise ValueError(
                    f"Number of headers ({len(headers)}) does not match "
                    f"specified number of columns ({min_col=}, {max_col=})"
                )
        self.curr_row = row + 1
        self.min_col = min_col
        self.max_col = self.min_col + len(headers) - 1
        return headers

    def __iter__(self):
        return self

    def __next__(self):
        min_row = self.curr_row
        min_col = self.min_col or 1
        max_col = self.max_col
        max_row = self.max_row or self.worksheet.max_row
        for row in self.worksheet.iter_rows(
            values_only=self.values_only,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            self.curr_row += 1
            if self.skip_blank_rows and not any(cell for cell in row):
                continue
            return dict(zip(self.headers, row))
        raise StopIteration
