import os.path

from openpyxl import load_workbook, Workbook


def read_excel(excel_path: str, sheet_name: str = None, header: int = 1):
    wb = load_workbook(excel_path)
    if sheet_name is None:
        sheet_names = wb.sheetnames
        sheet_data_list = []
        for name in sheet_names:
            tmp = []
            for row in wb[name].values:
                tmp.append(row)
            del tmp[:header]
            sheet_data_list.append(tmp)
        print(sheet_data_list)
        return sheet_data_list
    else:
        if sheet_name not in wb.sheetnames:
            error_info = f"{excel_path}表格中不存在: {sheet_name}，请核对sheet_name参数!"
            raise ValueError(error_info)
        sheet_name_data = []
        sheet = wb[sheet_name]
        for row in sheet.values:
            sheet_name_data.append(row)
        del sheet_name_data[:header]
        print(sheet_name_data)
        return sheet_name_data


def write_excel(excel_path: str, sheet_name: str, sheet_data=(), new=False):
    file_suffix = ["xlsx", "xlsm"]
    if os.path.exists(excel_path):
        print("文件已存在")
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()  # 新建工作簿

    if excel_path.split(".")[1] not in file_suffix:
        raise ValueError(f"文件名必须以: {file_suffix}结尾")
    if new:
        wb.create_sheet(sheet_name)
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    if sheet_name is not None:
        sheet.title = sheet_name
    for row in sheet_data:
        sheet.append(row)
    wb.save(excel_path)
    wb.close()


