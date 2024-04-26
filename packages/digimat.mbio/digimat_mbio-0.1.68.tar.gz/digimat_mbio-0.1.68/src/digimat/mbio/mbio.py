#!/bin/python

#      _ _       _                 _               _     _
#   __| (_) __ _(_)_ __ ___   __ _| |_   _ __ ___ | |__ (_) ___
#  / _` | |/ _` | | '_ ` _ \ / _` | __| | '_ ` _ \| '_ \| |/ _ \
# | (_| | | (_| | | | | | | | (_| | |_ _| | | | | | |_) | | (_) |
#  \__,_|_|\__, |_|_| |_| |_|\__,_|\__(_)_| |_| |_|_.__/|_|\___/
#          |___/                     The Digimat MBIO Processor


from __future__ import annotations
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from .values import MBIOValue

import pkg_resources

from threading import Thread, Lock
from queue import Queue

# import time
import logging
import logging.handlers
import xml.etree.ElementTree as ET
import signal
import os
import time
import threading

import base64

from prettytable import PrettyTable
import ipcalc

from .gateway import MBIOGateway, MBIOGatewayMetzConnect, MBIOGateways
from .task import MBIOTasks
from .xmlconfig import XMLConfig
from .valuenotifier import MBIOValueNotifier
from .linknotifier import MBIOTaskLinkNotifier


from .gateway import MBIOGatewayMetzConnectConfigurator


class MBIO(object):
    """MBIO processor base object"""
    def __init__(self, config='config.xml', logServer='localhost', logLevel=logging.DEBUG, interface=None):
        logger=logging.getLogger("MBIO")
        logger.setLevel(logLevel)
        socketHandler = logging.handlers.SocketHandler(logServer, logging.handlers.DEFAULT_TCP_LOGGING_PORT)
        logger.addHandler(socketHandler)
        self._logger=logger

        self._interface=interface
        self._network=None

        self._eventStop=threading.Event()
        self._threadManager=threading.Thread(target=self.manager)
        self._threadManager.daemon=True
        self._threadManager.start()

        self._gateways=MBIOGateways(logger)
        self._tasks=MBIOTasks(logger)
        self._valuesByKey: dict[str, MBIOValue]={}
        self._valuesWithManager: dict[str, MBIOValue]={}
        self._valueNotifiers=[]

        self.logger.info('Using MBIO v%s' % self.version)

        self._xmlConfigFilePath=config
        self._xmlConfig=None
        self.load(config)

        self.registerSignalHandlers()

    @property
    def logger(self):
        return self._logger

    @property
    def parent(self):
        return None

    def rootPath(self):
        try:
            return os.path.dirname(self._xmlConfigFilePath)
        except:
            pass

    def getVersion(self):
        """Return the version of the module"""
        try:
            distribution=pkg_resources.get_distribution('digimat.mbio')
            return distribution.parsed_version
        except:
            pass

    @property
    def version(self):
        """Version of the module"""
        return self.getVersion()

    @property
    def interface(self):
        """IP address of the MBIO network interface (declared in the config file)"""
        return self._interface

    @property
    def network(self):
        """Network (i.e 192.168.1.0/24) corresponding to the MBIO LAN interface (declared in the config file)"""
        return self._network

    def registerSignalHandlers(self):
        """register signal handlers"""
        def signalHandlerHALT(signum, frame):
            self.stop()
            try:
                exit(0)
            except:
                pass

        def signalHandlerIGNORE(signum, frame):
            pass

        self.logger.debug('Registering signal handlers')
        signal.signal(signal.SIGINT, signalHandlerHALT)
        signal.signal(signal.SIGHUP, signalHandlerHALT)
        signal.signal(signal.SIGTERM, signalHandlerHALT)
        signal.signal(signal.SIGUSR1, signalHandlerIGNORE)
        signal.signal(signal.SIGUSR2, signalHandlerIGNORE)

    def loadDevices(self, gateway, xml: XMLConfig):
        try:
            devices=xml.children('device')
            if devices:
                for d in devices:
                    if not d.getBool('enable', True):
                        continue

                    vendor=d.get('vendor')
                    model=d.get('model')
                    address=d.getInt('address')
                    try:
                        if gateway.device(address):
                            self.logger.error('Duplicate DEVICE declaration GW%s[%d]' % (gateway.key, address))
                            continue

                        gateway.declareDeviceFromName(vendor, model, address, xml=d)

                    except:
                        self.logger.exception('x')
                        self.logger.error('Error declaring DEVICE GW:%s (%s)' % (gateway.key, d.tostring()))

            item=xml.child('discover')
            if item:
                start=item.getInt('start', 1)
                end=item.getInt('end', 32)
                maxerrors=item.getInt('maxerrors', 3)
                gateway.discover(start, end, maxerrors)
        except:
            self.logger.exception('loadDevices(GW:%s)' % gateway.key)

    def loadGateways(self, xml: XMLConfig):
        try:
            gateways=xml.children('gateway')
            if gateways:
                for g in gateways:
                    if not g.getBool('enable', True):
                        continue

                    name=g.get('name')
                    host=g.get('host')
                    if name and host:
                        if self.gateway(host):
                            self.logger.error('Duplicate GATEWAY declaration %s:%s' % (name, host))
                            continue

                        gateway=self.declareGateway(name=name, host=host, port=g.get('port', 502), xml=g)
                        if gateway:
                            self.loadDevices(gateway, g)
                            continue

                    self.logger.error('Error declaring GATEWAY (%s)' % (g.tostring()))
        except:
            self.logger.exception('loadGateways')

    def load(self, fpath):
        try:
            root=ET.parse(fpath).getroot()

            # Make tree lowercase
            data=ET.tostring(root).lower()
            root=ET.fromstring(data)

            config=XMLConfig(root)
            self._interface=config.get('interface', self._interface)
            self._network=config.get('network')
            self.loadGateways(config)

            # TODO: better import design

            from .tasktest import MBIOTaskPulsar
            for item in config.children('pulsar'):
                MBIOTaskPulsar(self, item.get('name'), xml=item)

            from .tasktest import MBIOTaskCopier
            for item in config.children('copier'):
                MBIOTaskCopier(self, item.get('name'), xml=item)

            from .vio import MBIOTaskVirtualIO
            for item in config.children('virtualio'):
                MBIOTaskVirtualIO(self, item.get('name'), xml=item)

            from .mystrom import MBIOTaskMyStromSwitch
            for item in config.children('mystrom'):
                MBIOTaskMyStromSwitch(self, item.get('name'), xml=item)

            from .scheduler import MBIOTaskScheduler
            for item in config.children('scheduler'):
                MBIOTaskScheduler(self, item.get('name'), xml=item)

            from .meteosuisse import MBIOTaskMeteoSuisse
            for item in config.children('meteosuisse'):
                MBIOTaskMeteoSuisse(self, item.get('name'), xml=item)

            from .vplc import MBIOTaskVPLCTest
            for item in config.children('vplctest'):
                MBIOTaskVPLCTest(self, item.get('name'), xml=item)

            from .sensorpush import MBIOTaskSensorPush
            for item in config.children('sensorpush'):
                MBIOTaskSensorPush(self, item.get('name'), xml=item)

            from .hue import MBIOTaskHue
            for item in config.children('hue'):
                MBIOTaskHue(self, item.get('name'), xml=item)

            notifiers=config.child('notifiers')
            if notifiers:
                targets=notifiers.children('target')
                for target in targets:
                    self._cpuNotifier=MBIOTaskLinkNotifier(self, target.get('name'), xml=target)

                filters=notifiers.children('filter')
                if filters:
                    for f in filters:
                        for rule in f.children():
                            key=rule.get('key')
                            if key:
                                values=self.values(key)
                                if values:
                                    if rule.tag=='disable':
                                        for value in values:
                                            value.enableNotify(False)
                                    elif rule.tag=='enable':
                                        for value in values:
                                            value.enableNotify(True)

        except:
            self.logger.exception('xml')

    def RAZCPUCONFIG(self, proof, what=None):
        """Try to find a LinkNotifier and call RAZCONFIG (proof must be YESIAMSURE)"""
        try:
            if proof=="YESIAMSURE":
                return self._cpuNotifier.RAZCpuConfig(what)
        except:
            pass

    def beep(self):
        """Send a beep message request to the notifier (who should beep)"""
        self._cpuNotifier.beep()

    @property
    def tasks(self) -> MBIOTasks:
        """Registered tasks"""
        return self._tasks

    @property
    def gateways(self) -> MBIOGateways:
        """Registered gateways"""
        return self._gateways

    def gateway(self, gid):
        """Return the task given by it's id (key, name)"""
        return self._gateways.item(gid)

    def declareGateway(self, name, host, port=502, timeout=3, retries=3, xml: XMLConfig = None):
        gateway=self.gateway(host)
        if gateway:
            return gateway
        if name is None:
            name='%d' % len(self._gateways)
        gtype=xml.get('model')
        if gtype=='metzconnect':
            gateway=MBIOGatewayMetzConnect(self, name, host=host, port=port, interface=self._interface, timeout=timeout, retries=retries, xml=xml)
        else:
            gateway=MBIOGateway(self, name, host=host, port=port, interface=self._interface, timeout=timeout, retries=retries, xml=xml)
        self._gateways.add(gateway)
        return gateway

    def task(self, tid):
        return self._tasks.item(tid)

    def declareTask(self, task):
        self._tasks.add(task)

    def dump(self):
        t=PrettyTable()
        t.field_names=['Type', 'Key', 'State', 'Error']
        t.align='l'
        for gateway in self._gateways:
            state='CLOSED'
            if gateway.isOpen():
                state='OPEN'
            address='%s:%s' % (gateway.key, gateway.host)
            t.add_row([gateway.__class__.__name__, address, state, gateway.isError()])
        for task in self._tasks:
            t.add_row([task.__class__.__name__, task.key, task.statestr(), task.isError()])
        print(t)

    def dumpValues(self, key='*'):
        values=self.values(key)
        if values:
            t=PrettyTable()
            t.field_names=['Key', 'Value']
            t.align='l'

            for value in values:
                t.add_row([value.key, str(value)])

            print(t.get_string())

    def tree(self, key=None, values=False):
        from treelib import Tree

        tree=Tree()
        tree.create_node('%s' % str(self), "root")
        tree.create_node('Gateways', 'gateways', parent='root')
        tree.create_node('Tasks', 'tasks', parent='root')

        for gateway in self._gateways:
            label=str(gateway)
            tree.create_node(label, gateway.key, parent='gateways')
            for device in gateway.devices:
                label=str(device)
                tree.create_node(label, device.key, parent=gateway.key)
                for value in device.values:
                    if not value.match(key):
                        continue
                    label=str(value)
                    tree.create_node(label, value.key, parent=device.key)
                for value in device.sysvalues:
                    if not value.match(key):
                        continue
                    label=str(value)
                    tree.create_node(label, value.key, parent=device.key)

        for item in self._tasks:
            tree.create_node(str(item), item.key, parent='tasks')
            for value in item.values:
                if not value.match(key):
                    continue
                label=str(value)
                tree.create_node(label, value.key, parent=item.key)

        if values or key:
            tree.create_node('Values', 'values', parent='root')
            for item in self._valuesByKey.values():
                if not item.match(key):
                    continue
                tree.create_node(str(item), 'REF-'+item.key, parent='values')

        print(str(tree))

    def iterValuesWithManager(self):
        if self._valuesWithManager:
            try:
                return iter(self._valuesWithManager.values())
            except:
                pass

    def manager(self):
        self.logger.info('MBIO Manager Thread started')
        idle=True
        iterValues=None
        while not self._eventStop.isSet():
            # self.logger.debug('MANAGER')
            time.sleep(0)
            idle=True
            try:
                value=iterValues.next()
                value.manager()
                idle=False
            except:
                iterValues=self.iterValuesWithManager()

            if idle:
                time.sleep(1)

        self.logger.info('MBIO Manager Thread halted')

    def stop(self, code=0):
        """Stop the MBIO processor. Shoud be called before exiting to shutdown tasks"""
        self.logger.info('STOP!')
        self._eventStop.set()
        self._tasks.stop()
        self._tasks.waitForThreadTermination()
        self._gateways.stop()
        self._gateways.waitForThreadTermination()
        self._threadManager.join()
        try:
            exit(code)
        except:
            pass
        try:
            quit()
        except:
            pass

    def reset(self):
        """Reset (restart) all tasks and gateways"""
        self._tasks.reset()
        self._gateways.reset()

    def resetHalted(self):
        """Reset (restart) all tasks and gateways that are halted"""
        self._tasks.resetHalted()
        self._gateways.resetHalted()

    def halt(self):
        """Halt (pause) all tasks and gateways. A restart is done with reset()"""
        self._tasks.halt()
        self._gateways.halt()

    def discover(self):
        """Start a device discover on all gateways"""
        self._gateways.discover()

    def __getitem__(self, key):
        """Return the task, gateway or value given by it's key"""
        item=self.gateway(key)
        if item:
            return item
        item=self.task(key)
        if item:
            return item
        return self.value(key)

    def __del__(self):
        self.stop()

    def __repr__(self):
        return '%s(%d gateways, %d tasks, %d values, %d notifiers)' % (self.__class__.__name__,
            len(self._gateways),
            len(self._tasks),
            len(self._valuesByKey),
            len(self._valueNotifiers))

    def registerValue(self, value):
        """Any MBIOValue added to a MBIOValues collection will call this function to register itself to the MBIO"""
        # assert(isinstance(value, MBIOValue))
        if value is not None:
            self._valuesByKey[value.key]=value
            if value.hasManager():
                self._valuesWithManager[value.key]=value

    def value(self, key):
        """Return the MBIOValue given by it's key"""
        try:
            return self._valuesByKey[key]
        except:
            pass

    def values(self, key='*', sort=True):
        """Return the MBIOValues given by it's key using value.match() """
        values=[]
        for value in self._valuesByKey.values():
            if value.match(key):
                values.append(value)
        if sort:
            values=sorted(values, key=lambda value: value.key)
        return values

    def topvalues(self, key='*'):
        """Return the MBIOValues given by it's key using value.match() sorted by number of notifyCount"""
        values=[]
        for value in self._valuesByKey.values():
            if value.match(key):
                values.append(value)
        values=sorted(values, key=lambda value: value._notifyCount, reverse=True)
        return values

    def errorvalues(self, key='*'):
        """Return the MBIOValues with error flag active"""
        values=[]
        for value in self._valuesByKey.values():
            if value.isError() and value.match(key):
                values.append(value)
        return values

    def registerValueNotifier(self, notifier: MBIOValueNotifier):
        """Register a MBIOValueNotifier (receive every value notification update)"""
        assert(isinstance(notifier, MBIOValueNotifier))
        self._valueNotifiers.append(notifier)

    def signalValueUpdateNotify(self, value):
        """MBIOValue update notification that will be sent to every registered notifier"""
        if value is not None and self._valueNotifiers and value._enableNotify:
            if value.isEnabled():
                for notifier in self._valueNotifiers:
                    notifier.put(value)

    def renotifyValues(self):
        """Trigger an update notify for every registered MBIOValue"""
        if self._valuesByKey:
            for value in self._valuesByKey.values():
                value.notify(force=True)

    def on(self):
        """Send an ON to every registerred MBIOValue. Just another debug tool"""
        for gateway in self.gateways:
            gateway.on()

    def off(self):
        """Send an ON to every registerred MBIOValue. Just another debug tool"""
        for gateway in self.gateways:
            gateway.off()

    def toggle(self):
        """Send a TOGGLE to every registerred MBIOValue. Just another debug tool"""
        for gateway in self.gateways:
            gateway.toggle()

    def exportValuesToXLSX(self, fpath, showref=False):
        """Create a XLSX file containing a list of MBIO values"""
        from openpyxl import Workbook
        wb=Workbook()
        ws=wb.active
        ws.title='MBIO'

        headers=['KEY', 'TYPE', 'VALUE', 'UNIT', 'FLAGS']
        if showref:
            headers.append('REF')

        col=1
        for data in headers:
            ws.cell(1, col).value=data
            col+=1

        li=2
        for value in self.values():
            ws.cell(li, 1).value=value.key
            ws.cell(li, 2).value=value.type
            ws.cell(li, 3).value=value.value
            ws.cell(li, 4).value=value.unitstr()
            ws.cell(li, 5).value=value.flags
            if showref:
                ws.cell(li, 6).value=value.refpath()
            li+=1

        wb.save(fpath)

    def scanbus(self, network=None):
        """"""
        network=network or self.network
        scanner=MBIOGatewayMetzConnectScanner(self, network)
        if scanner.discover():
            self.stop()

    def b16encode(self, s):
        try:
            return '(b16)%s' % base64.b16encode(s.encode()).decode()
        except:
            pass


class MBIONetworkScanner(object):
    def __init__(self, mbio: MBIO, network=None, maxthreads=128):
        self._mbio=mbio
        self._network=network or mbio._network
        self._maxthreads=maxthreads
        self._queue=Queue()
        self._lock=Lock()
        self._hosts=[]
        self.onInit()

    def onInit(self):
        pass

    @property
    def mbio(self):
        return self._mbio

    @property
    def logger(self):
        return self.mbio.logger

    def hosts(self):
        hosts=[]
        try:
            n=ipcalc.Network(self._network)
            for host in n:
                hosts.append(str(host))
        except:
            pass
        return hosts

    def probe(self, host):
        # to be overriden
        return False

    def probeNetwork(self):
        while True:
            try:
                host=self._queue.get()
            except:
                break

            try:
                self.logger.debug('%s: scanning host %s...' % (self.__class__.__name__, host))
                if self.probe(host):
                    with self._lock:
                        self.logger.warning('%s: found host %s...' % (self.__class__.__name__, host))
                        self._hosts.append(host)
            except:
                pass
            self._queue.task_done()

    def scan(self):
        if self._network:
            hosts=self.hosts()
            if hosts:
                for t in range(self._maxthreads):
                    t = Thread(target=self.probeNetwork)
                    t.daemon = True
                    t.start()

                for host in hosts:
                    self._queue.put(host)

                self._queue.join()
                return self._hosts


class MBIOGatewayMetzConnectScanner(MBIONetworkScanner):
    def onInit(self):
        self._gateways={}

    def configurator(self, mac):
        if mac:
            host=self._gateways.get(mac.lower())
            if host:
                c=MBIOGatewayMetzConnectConfigurator(host, None)
                return c

    def probe(self, host):
        result=False
        try:
            self.logger.debug('Scanning host %s for MetzConnect GW...' % host)
            c=MBIOGatewayMetzConnectConfigurator(host, None)
            if c.connect():
                if c['mac']:
                    with self._lock:
                        self.logger.warning('Found MetzConnect %s gateway @%s' % (c['mac'], host))
                        result=True
                        self._gateways[c['mac'].lower()]=host
            c.disconnect()
            del(c)
        except:
            self.logger.exception('probe')
            pass
        return result

    def discover(self):
        reconfigured=[]
        gateways=[]
        for gateway in self.mbio.gateways:
            if gateway.model=='metzconnect' and gateway.MAC:
                gateways.append(gateway)

        if gateways and self.scan():
            for gateway in gateways:
                ip=self._gateways.get(gateway.MAC)
                if gateway.host != ip:
                    self.logger.warning('MetzConnect Gateway %s/%s should be %s' % (gateway.MAC, ip, gateway.host))
                    c=self.configurator(gateway.MAC)
                    if c.connect():
                        self.logger.warning('Reconfiguring MetzConnect Gateway %s...' % (gateway.MAC))
                        c.setIP(gateway.host)
                        c.name=gateway.name
                        c.setRS485(19200, 'E', 1, 1.0, True)
                        c.idletimeout=30
                        c.password=gateway.password
                        c.netrestart()
                        self.logger.warning('Restarting Gateway %s network...' % (gateway.MAC))
                        reconfigured.append(gateway)
                    c.disconnect()
        return reconfigured


if __name__=='__main__':
    pass
